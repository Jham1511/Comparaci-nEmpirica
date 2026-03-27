import time
import statistics
import openpyxl

def factorial_iterativo(n):
    resultado = 1
    for i in range(1, n + 1):
        resultado *= i
    return resultado

def factorial_recursivo(n):
    if n <= 1:
        return 1
    return n * factorial_recursivo(n - 1)

def medir_tiempos(func, n, repeticiones=20):
    tiempos = []
    for _ in range(repeticiones):
        inicio = time.perf_counter()
        func(n)
        fin = time.perf_counter()
        tiempos.append((fin - inicio) * 1000)
    return tiempos

valores_n = [10, 20, 30, 40, 50, 60, 70, 80, 90, 100, 200, 300, 400, 500, 600, 700]

wb = openpyxl.load_workbook('Plantilla Mediciones.xlsx')

for n in valores_n:
    tiempos_iter = medir_tiempos(factorial_iterativo, n)
    tiempos_rec = medir_tiempos(factorial_recursivo, n)
    
    media_iter = statistics.mean(tiempos_iter)
    media_rec = statistics.mean(tiempos_rec)
    desv_iter = statistics.stdev(tiempos_iter)
    desv_rec = statistics.stdev(tiempos_rec)
    
    ws = wb.copy_worksheet(wb['Plantilla Mediciones'])
    ws.title = f'n={n}'
    
    ws['B6'] = f'Input n = {n}'
    
    for i in range(20):
        ws.cell(row=8 + i, column=4, value=tiempos_iter[i])
        ws.cell(row=8 + i, column=6, value=tiempos_rec[i])
    
    ws['D28'] = media_iter
    ws['F28'] = media_rec
    ws['D29'] = desv_iter
    ws['F29'] = desv_rec

wb.remove(wb['Plantilla Mediciones'])
wb.save('Plantilla Mediciones_completada.xlsx')
print("Archivo generado: Plantilla Mediciones_completada.xlsx")